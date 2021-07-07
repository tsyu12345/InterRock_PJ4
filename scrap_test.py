from base64 import standard_b64decode
from PySimpleGUI.PySimpleGUI import No
import openpyxl as px
import PySimpleGUI as gui
import re
import sys
import os
import threading as th
import time
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, WebDriverException
#from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup as bs


class Scrap():
    book = px.Workbook()
    sheet = book.worksheets[0]

    def __init__(self, path):
        # init WebDriver
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_argument("enable-automation")
        # options.add_argument("--headless")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument('--disable-extensions')
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-browser-side-navigation")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument('--ignore-certificate-errors')
        self.options.add_argument('--ignore-ssl-errors')
        prefs = {"profile.default_content_setting_values.notifications": 2}
        self.options.add_experimental_option("prefs", prefs)
        self.driver = webdriver.Chrome(
            executable_path='chromedriver.exe', options=self.options)
        self.path = path

    def book_init(self):
        col_list = [
            "エキテン",
            "新規リスト投入日",
            "ジャンル",
            "電話",
            "店舗名",
            "店舗名カナ",
            "料金プラン",
            "都道府県コード",
            "都道府県",
            "市区町村・番地",
            "住所フル",
            "店舗URL",
            "shopID",
            "URLその1",
            "URLその2",
            "URLその3",
            "パンくず",
            "キャッチ",
            "店舗公式",
            "未確認店舗",
            "評価点数",
            "口コミ数",
            "写真枚数",
            "アクセス",
            "ジャンル1",
            "ジャンル2",
            "ジャンル3",
            "ジャンル4",
            "早朝OK",
            "日祝OK",
            "夜間OK",
            "駐車場有",
            "ネット予約",
            "クーポン有",
            "カード可",
            "出張・宅配あり",
            "緯度",
            "経度",
            "アクセス／最寄駅",
            "営業時間／定休日",
            "駐車場",
            "クレジットカード",
            "座席",
            "用途",
            "メニュー",
            "特徴",
            "ポイント",
            "ここがすごい！",
            "メディア関連",
            "価格設定",
            "マルチアクセス",
            "紹介文"
        ]
        for col, menu in enumerate(col_list):
            self.sheet.cell(row=1, column=col+1, value=menu)
        self.book.save(self.path)

    def junle_list(self):
        list = [
            "リラク・ボディケア",
            "ヘアサロン・ネイル",
            "学習塾・予備校",
            "習い事・スクール",
            "歯科・矯正歯科",
            "医院・クリニック・ヘルスケア",
            "ショッピング",
            "お出かけ・レジャー",
            "リサイクル・中古買取り",
            "ペット・動物",
            "出張デリバリー・生活サービス",
            "住宅・不動産",
            "冠婚葬祭"
        ]
        return list

    def search(self, area):  # 検索と条件指定
        #driver_action = ActionChains(self.driver)
        self.driver.get('https://www.ekiten.jp/')
        sr_box = self.driver.find_element_by_css_selector(
            '#select_form_st_com')
        sr_box.send_keys(area)
        sr_btn = self.driver.find_element_by_css_selector(
            '#js_random_top > div > div > div > form > div > input')
        sr_btn.click()
        city_list = self.extraction_url(
            'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > div.grouped_list_body > ul > li > a', 'https://www.ekiten.jp')

        for city in city_list:
            self.driver.get(city)
            print(city)
            time.sleep(1)
            select = self.driver.find_element_by_css_selector(
                'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > a').text

            # 区町村選択がない場合の処理系
            if select in '駅・バス停から探す ':
                station_list = self.extraction_url(
                    '#tab_point_0 > div > div > ul > li > a', 'https://www.ekiten.jp')
                print(station_list)
                # another process here
                for station in station_list:
                    self.driver.get(station)
                    time.sleep(1)
                    junle_list = self.extraction_url(
                        'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                    print(junle_list)
                    for junle in junle_list:
                        self.driver.get(junle)
                        time.sleep(1)
                        kategoli_list = self.extraction_url(
                            'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                        print(kategoli_list)
                        for kategoli in kategoli_list:
                            self.driver.get(kategoli)
                            # scrap URL process here
                            self.scrap_url()
                        time.sleep(3)
                    self.restart()

            # 区町村選択がある場合の処理系
            else:
                city_list2 = self.extraction_url(
                    'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                for city2 in city_list2:
                    self.driver.get(city2)
                    time.sleep(1)
                    station_list = self.extraction_url(
                        '#tab_point_0 > div > div > ul > li > a', 'https://www.ekiten.jp')
                    if station_list == []:
                        station_list = self.extraction_url(
                            '#tab_point_1 > div > div > ul > li > a', 'https://www.ekiten.jp')
                    print(station_list)
                    for station in station_list:
                        self.driver.get(station)
                        time.sleep(1)
                        junle_list = self.extraction_url(
                            'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                        print(junle_list)
                        for junle in junle_list:
                            self.driver.get(junle)
                            time.sleep(1)
                            kategoli_list = self.extraction_url(
                                'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                            print(kategoli_list)
                            for kategoli in kategoli_list:
                                self.driver.get(kategoli)
                                self.scrap_url()
                                # scrap URL process here
                            time.sleep(3)
                        time.sleep(3)
                    self.restart()
            self.restart()

    def restart(self):
        self.driver.quit()
        time.sleep(3)
        self.driver = webdriver.Chrome(
            'chromedriver.exe', options=self.options)

    def scrap_url(self):
        while True:
            html = self.driver.page_source
            soup = bs(html, 'lxml')
            try:
                a_tags = soup.select('.p-shop_box_head_title_body > a')
                print(a_tags)
                self.write_url(a_tags)
                next_btn = self.driver.find_element_by_css_selector(
                    'div.p-pagination_next > a')
                next_btn.click()
            except NoSuchElementException:
                break

    def info_scrap(self, url):
        write_data = [] #書き込み用データを保存するリスト：各要素はブックの列に対応
        self.driver.get(url)
        html = self.driver.page_source
        soup = bs(html, 'lxml')
        write_data.append(
            soup.select_one(
                'div.layout_media p-topic_path_container > div.topic_path > a:nth-child(3)'
            ).get_text()
        )#ジャンル
        try:
            tel = soup.select_one(
                    'body > div:nth-child(12) > div > div.p-tel_modal_phone_number > div > div > p'
                ).get_text()
        except AttributeError:
            tel = None
        write_data.append(tel)#TEL
        # table info scraping
        table_col = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.p-shop_content_container.p-shop_content_container_relative > table > tbody > tr > th'
            )
        table_info = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.p-shop_content_container.p-shop_content_container_relative > table > tbody > tr >td'
        )
        table_list = {}
        for menu, info in zip(table_col, table_info):
            menu = menu.get_text()
            info = info.get_text()
            info = info.replace(" ", "")
            info = info.replace("　", "")
            info = info.replace("\n", "")
            info = info.replace("地図で場所を見るGoogleマップで見る")
            table_list[menu] = info
        write_data.append(table_list['店舗名'])
        try:
            store_kana = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div > div.p-shop_header_name_container > div > span').get_text()
        except AttributeError:
            store_kana = None
        write_data.append(store_kana)
        all_addresses = table_list['住所']
        pref_obj = re.search('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_addresses)
        pref = pref_obj.group()
        write_data.append(self.call_jis_code(pref)) #JISコード
        write_data.append(pref) #都道府県名
        muni = re.split('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_addresses)#都道府県と市区町村を分離
        write_data.append(muni[1]) #市区町村番地
        write_data.append(all_addresses)#フル住所
        write_data.append(url) #店舗URL
        shop_id_string = re.search(r"shop_\d{0,}", url).group()
        shop_id = re.search(r"\d{1,}", shop_id_string).group()
        write_data.append(shop_id) #店舗ID
        url_list = re.findall(r"https?://[\w!\?/\+\-_~=;\.,\*&@#\$%\(\)'\[\]]+", table_list['URL'])
        for i in range(3):
            write_data.append(url_list[i]) #URLその１～３        
        pankuzu_header = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.layout_media.p-topic_path_container > div.layout_media_wide > div').get_text()
        pankuzu = pankuzu_header.replace('\n', " > ")
        write_data.append(pankuzu.strip(" > ")) #パンくずヘッダー
        try:
            catch_copy = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.p-shop_header_catch_container > p').get_text()
        except AttributeError:
            catch_copy = None
        write_data.append(catch_copy)
        official = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.p-shop_header_catch_container > div > span')
        if official != None:
            write_data.append("●") #店舗公式
        else:
            write_data.append(None)
        write_data.append(None) #未確認店舗
        store_score_tag = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_rating > div > div.rating_stars_num.tooltip > span')
        if store_score_tag == None:
            write_data.append(None) #評価点数
        else:
            score = store_score_tag.get_text()
            write_data.append(score)
        


        #wirte book
        for col, data in enumerate(write_data):
            self.sheet.cell(row=self.sheet.max_row + 1, column=col+1, value=data)
        self.book.save(self.path)
        
    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": 1,
            "青森県": 2,
            "岩手県": 3,
            "宮城県": 4,
            "秋田県": 5,
            "山形県": 6,
            "福島県": 7,
            "茨城県": 8,
            "栃木県": 9,
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        return code

    def write_url(self, a_tag_list):
        url_list = []
        for a in a_tag_list:
            url_list.append("https://www.ekiten.jp" + a.get('href'))
        for url in url_list:
            row = self.sheet.max_row + 1
            print(row)
            self.sheet.cell(row=row, column=13, value=url)
        self.book.save(self.path)

    def extraction_url(self, selector, pre_url):
        html = self.driver.page_source
        soup = bs(html, 'lxml')
        s = soup.select(selector)
        # print(s)
        url_list = self.return_url(s, pre_url)
        return url_list

    def return_url(self, a_tag_list, pre_url):
        url_list = []
        for a in a_tag_list:
            url_list.append(pre_url+a.get('href'))
        return url_list


if __name__ == "__main__":
    scraping = Scrap('./test.xlsx')
    scraping.book_init()
    scraping.search('徳島県')
    # scraping.scrap()
    scraping.driver.quit()
