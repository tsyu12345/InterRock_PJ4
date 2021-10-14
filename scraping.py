from multiprocessing.managers import RemoteError
from PySimpleGUI.PySimpleGUI import T, No
from bs4.builder import TreeBuilder
import openpyxl as px
import re
import sys
import os
import time
from requests import api
from selenium import webdriver
from selenium.common.exceptions import InvalidSessionIdException, NoSuchElementException, WebDriverException, TimeoutException
#from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup as bs
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import requests
from multiprocessing import Pool, Manager

class ScrapingURL(object):
    
    def __init__(self, path, row_counter, sync_data_list):
        self.path = path
        self.driver_path = 'chromedriver.exe'
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_argument("enable-automation")
        self.options.add_argument("--headless")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument('--disable-extensions')
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-browser-side-navigation")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument('--ignore-certificate-errors')
        self.options.add_argument('--ignore-ssl-errors')
        prefs = {"profile.default_content_setting_values.notifications": 2}
        self.options.add_experimental_option("prefs", prefs)
        browser_path = 'chrome-win/chrome.exe'
        self.options.binary_location = browser_path
        #self.driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        self.writeRow = 0
        self.row_counter = row_counter
        self.url_list = sync_data_list

    def search(self, area_list:list):
        for area in area_list:
            self.__searchOperation(area)

    # 全ジャンル抽出の場合
    def __searchOperation(self, area):  # 検索と条件指定
        """
        指定エリア（1県分）のサイトURLを検索する。
        """
        #driver_action = ActionChains(self.driver)
        # self.driver.implicitly_wait(5)
        self.sub_driver = webdriver.Chrome('chromedriver.exe', options=self.options)
        self.sub_driver.get('https://www.ekiten.jp/')
        # wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#select_form_st_com")))
        # Max wait time(second):180s
        self.wait_sub_driver = WebDriverWait(self.sub_driver, 180)
        #self.wait_sub_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'select_form_st_com')))
        sr_box = self.sub_driver.find_element_by_id('select_form_st_com')
        sr_box.send_keys(area)
        sr_btn = self.sub_driver.find_element_by_css_selector(
            '#js_random_top > div > div > div > form > div > input')
        sr_btn.click()
        print("clicked !")
        self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
        city_list = self.extraction_url(
            'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > div.grouped_list_body > ul > li > a', 'https://www.ekiten.jp')
        print(city_list)
        result:str = self.sub_driver.find_element_by_css_selector(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.search_result_heading.u-mb10 > div.search_result_heading_sub > dl > div > dd').text
        result = result.replace(",", "")
        result = result.replace("件", "")
        self.result_cnt = int(result)
        self.count = 0
        for city in city_list:
            self.sub_driver.get(city)
            # wait.until(EC.visibility_of_all_elements_located)
            # print(city)
            # time.sleep(1)
            # WebDriverWait(self.driver, 180)#Max wait time(second):180s
            self.wait_sub_driver.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > a')))
            select = self.sub_driver.find_element_by_css_selector(
                'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > a').text

            # 区町村選択がない場合の処理系
            if select in '駅・バス停から探す ':
                # time.sleep(1)
                self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                junle_list = self.extraction_url(
                    'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                # print(junle_list)
                for junle in junle_list:
                    print(junle)
                    self.sub_driver.get(junle)
                    self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                    # time.sleep(1)
                    kategoli_list = self.extraction_url(
                        'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                    # print(kategoli_list)
                    for kategoli in kategoli_list:
                        print(kategoli)
                        self.sub_driver.get(kategoli)
                        self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                        # scrap URL process here
                        self.scrap_url()
                        self.count += 1
                    self.__sub_restart()
                self.__sub_restart()

            # 区町村選択がある場合の処理系
            else:
                self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                city_list2 = self.extraction_url(
                    'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                for city2 in city_list2:
                    print(city2)
                    self.sub_driver.get(city2)
                    self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                    # time.sleep(1)
                    junle_list = self.extraction_url(
                        'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                    # print(junle_list)
                    for junle in junle_list:
                        print(junle)
                        self.sub_driver.get(junle)
                        self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                        # time.sleep(1)
                        kategoli_list = self.extraction_url(
                            'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                        # print(kategoli_list)
                        for kategoli in kategoli_list:
                            print(kategoli)
                            self.sub_driver.get(kategoli)
                            self.wait_sub_driver.until(
                                EC.visibility_of_all_elements_located)
                            self.scrap_url()
                            self.count += 1
                            # scrap URL process here
                        self.__sub_restart()
                self.__sub_restart()
            self.__sub_restart()
        self.__sub_restart()

    def __sub_restart(self):
        self.sub_driver.delete_all_cookies()
        self.sub_driver.quit()
        time.sleep(3)
        self.sub_driver = webdriver.Chrome(
            'chromedriver.exe', options=self.options)
        self.wait_sub_driver = WebDriverWait(self.sub_driver, 180)

    def scrap_url(self):
        while True:
            html = self.sub_driver.page_source
            soup = bs(html, 'lxml')
            try:
                a_tags = soup.select('.p-shop_box_head_title_body > a')
                # print(a_tags)
                self.write_url(a_tags)
                next_btn = self.sub_driver.find_element_by_css_selector(
                    'div.p-pagination_next > a')
                next_btn.click()
            except NoSuchElementException:
                break
    
    def write_url(self, a_tag_list):
        """
        抽出URLの書き込み(共有メモリ内配列への書き込み)
        URLを重複なくリストへ格納する。
        """
        url_pre_list = []
        for a in a_tag_list:
            url_pre_list.append("https://www.ekiten.jp" + a.get('href'))
            
        for url in url_pre_list:
            if url not in self.url_list:
                self.url_list.append(url)
                self.row_counter.value = len(self.url_list)
           
           
    def extraction_url(self, selector, pre_url):
        html = self.sub_driver.page_source
        soup = bs(html, 'lxml')
        s = soup.select(selector)
        # #print(s)
        url_list = self.return_url(s, pre_url)
        return url_list

    def return_url(self, a_tag_list, pre_url):
        url_list = []
        for a in a_tag_list:
            url_list.append(pre_url+a.get('href'))
        return url_list
    
class ScrapingInfomation(ScrapingURL):
    
    
    def __init__(self, path, row_counter, url_list_data, info_datas, end_count):
        super(ScrapingInfomation, self).__init__(path, row_counter, url_list_data)
        #self.driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        self.info_datas = info_datas #共有メモリ上の結果格納リスト
        self.end_count = end_count #共有メモリ上のカウンタ変数

    
    def requestHTML(self, url_list:list):
        """
        指定ページにアクセスしHTMLをロードする。
        その後、__extraction()を呼びだす。
        """
        driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        wait = WebDriverWait(driver, 180)
        load_counter = 0
        for url in url_list:
            if load_counter % 100 == 0 and load_counter != 0:
                driver.delete_all_cookies()
                driver.quit()
                time.sleep(10)
                driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
                wait = WebDriverWait(driver, 180) 
            
            try:
                driver.get(url)
            except TimeoutException:
                self.restart()
                time.sleep(30)
                driver.get(url)
                
            html = driver.page_source
            data_list:list = self.__extraction(html, url)
            self.info_datas.append(data_list) #結果を共有メモリ上のリストへ格納
            self.end_count.value += 1
            load_counter += 1
        driver.quit()
    
    def __create_data_list(self, data_length):
        """
        指定数分のNone要素のみのリストを作成。\n
        data = [None, None, None ......(len() = data_length)]
        """
        data_list = []
        for i in range(data_length):
            data_list.append(None)
        return data_list

    def __extraction(self, html, url) -> list:
        """
        HTMLを受け取り、結果をリストで返す。
        """
        data_list = self.__create_data_list(53)
        #data_list[0],[1] = None : A列とB列は空欄
        soup = bs(html, 'lxml')
        junle_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.layout_media.p-topic_path_container > div.layout_media_wide > div > a:nth-child(3)'
        )
        junle = junle_elm.get_text() if junle_elm != None else None
        data_list[2] = junle
        #self.sheet.cell(row=index, column=3, value=junle)  # ジャンル
        tel_elm = soup.select_one(
            'body > div:nth-child(12) > div > div.p-tel_modal_phone_number > div > div > p'
        )
        tel = tel_elm.get_text() if tel_elm != None else None
        data_list[3] = tel
        #self.sheet.cell(row=index, column=4, value=tel)  # TEL
        # table info scraping
        table_col = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.p-shop_content_container.p-shop_content_container_relative > table > tbody > tr > th'
        )
        table_info = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.p-shop_content_container.p-shop_content_container_relative > table > tbody > tr >td'
        )
        table_list = {}
        for menu, info in zip(table_col, table_info):
            try:
                menu = menu.get_text()
                info = info.get_text()
            except AttributeError:
                break
            info = info.replace(" ", "")
            info = info.replace("　", "")
            info = info.replace("\n", "")
            info = info.replace("地図で場所を見るGoogleマップで見る", "")
            table_list[menu] = info
        data_list[4] = table_list['店舗名']
        #self.sheet.cell(row=index, column=5, value=table_list['店舗名'])
        store_kana_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div > div.p-shop_header_name_container > div > span'
            )
        store_kana = store_kana_elm.get_text() if store_kana_elm != None else None
        data_list[5] = store_kana
        #self.sheet.cell(row=index, column=6, value=store_kana)
        all_addresses = table_list['住所']
        pref_obj = re.search('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_addresses)
        pref = pref_obj.group()
        data_list[7] = self.call_jis_code(pref)
        #self.sheet.cell(row=index, column=8, value=self.call_jis_code(pref))  # JISコード
        data_list[8] = pref
        #self.sheet.cell(row=index, column=9, value=pref)  # 都道府県名
        # 都道府県と市区町村を分離
        muni = re.split('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_addresses)
        data_list[9] = muni[1]
        data_list[10] = all_addresses
        data_list[11] = url
        """
        self.sheet.cell(row=index, column=10, value=muni[1])  # 市区町村番地
        self.sheet.cell(row=index, column=11, value=all_addresses)  # フル住所
        self.sheet.cell(row=index, column=12, value=url)  # 店舗URL
        """
        shop_id_string = re.search(r"shop_\d{0,}", url).group()
        shop_id = re.search(r"\d{1,}", shop_id_string).group()
        data_list[12] = shop_id
        #self.sheet.cell(row=index, column=13, value=shop_id)  # shopID
        try:
            url_list = re.findall(
                r"https?://[\w!\?/\+\-_~=;\.,\*&@#\$%\(\)'\[\]]+", table_list['URL'])
        except KeyError:  # URLがない時
            url_list = [None, None, None]
        for i in range(3):
            try:
                data_list[13+i] = url_list[i]
                #self.sheet.cell(row=index, column=14+i,value=url_list[i])  # URLその１～３
            except IndexError:
                data_list[13+i] = None
                #self.sheet.cell(row=index, column=14+i, value=None)
        pankuzu_header = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.layout_media.p-topic_path_container > div.layout_media_wide > div').get_text()
        pankuzu = pankuzu_header.replace('\n', " > ")
        data_list[16] = pankuzu.strip(" > ")
        #self.sheet.cell(row=index, column=17,value=pankuzu.strip(" > "))  # パンくずヘッダー
        catch_copy_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.p-shop_header_catch_container > p'
            )
        catch_copy = catch_copy_elm.get_text() if catch_copy_elm != None else None
        data_list[17] = catch_copy
        #self.sheet.cell(row=index, column=18, value=catch_copy)
        official = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.p-shop_header_catch_container > div > span')
        official_judge = '●' if official != None else None
        data_list[18] = official_judge
        #self.sheet.cell(row=index, column=19, value=official_judge)
        #self.sheet.cell(row=index, column=20, value=None)  # 未確認店舗
        store_score_tag = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_rating > div > div.rating_stars_num.tooltip > span'
            )
        if store_score_tag == None:
            score = None  # 評価点数
        else:
            score = float(store_score_tag.get_text())
        data_list[20] = score
        #self.sheet.cell(row=index, column=21, value=score)
        review_count = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_info > div.p-shop_header_info_review > div > span.icon_wrapper_text > a')
        if review_count in (None, '0'):
            review_count = 0
        else:
            review_count = int(review_count.get_text())
        data_list[21] = review_count
        #self.sheet.cell(row=index, column=22, value=review_count)  # 口コミ数
        photo_tag = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.l-shop_content > div:nth-child(3) > div.grid.space15.vertical_space15.js_photo_gallery > div > a > img')
        photo_cnt = len(photo_tag)  # 写真枚数
        data_list[22] = photo_cnt
        #self.sheet.cell(row=index, column=23, value=photo_cnt)
        access_elm = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_access > div:nth-child(1) > div')
        access = access_elm.get_text() if access_elm != None else None
        data_list[23] = access
        #self.sheet.cell(row=index, column=24, value=access)  # アクセス
        junle1_elm = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > ul.p-shop_header_genre > li > a')
        junle1 = junle1_elm.get_text() if junle1_elm != None else None
        data_list[24] = junle1
        #self.sheet.cell(row=index, column=25, value=junle1)  # ジャンル１
        junle2_4 = soup.select('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > ul.p-shop_header_genre > li > span')
        #ジャンル2~4
        for c, junle in enumerate(junle2_4):
            if junle != None:
                data_list[25+c] = junle.get_text()
                #self.sheet.cell(row=index, column=26+c, value=junle.get_text())
        
        
        features = soup.select('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > ul.p-shop_header_tag_list.tag_group > li')
        col_list = [
            '早朝OK',	
            '日祝OK',
            '夜間OK',
            '駐車場有',
            'ネット予約',	
            'クーポン有',
            'カード可',
            '出張・宅配あり'
        ]
        for feature in features:
            if feature == None:  # 特徴タグがない時処理を行わない
                break
            
            for col, col_menu in zip(range(29, 36+1), col_list): 
                if col_menu in feature.get_text():
                    data_list[28+i] = "●"
                    #self.sheet.cell(row=index, column=29+i, value="●")  # 店舗の特徴
                    break  # 見つかったら小ループ抜けて次の特徴へ
        
        #緯度・経度
        time.sleep(5) #API接続の感覚を空ける。
        latlong = self.calcLatLong(all_addresses)
        data_list[36] = latlong[0]
        data_list[37] = latlong[1]
        #self.sheet.cell(row=index, column=37, value=latlong[0])
        #self.sheet.cell(row=index, column=38, value=latlong[1])

        moyorieki = None
        data_list[38] = moyorieki
        #self.sheet.cell(row=index, column=39, value=moyorieki)  # アクセス/最寄り駅
        holiday_d_time_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_access > div:nth-child(3) > div'
        )
        holiday_d_time = holiday_d_time_elm.get_text() if holiday_d_time_elm != None else None
        try:
            holiday_d_time = holiday_d_time.replace("\n", "/")
            holiday_d_time = holiday_d_time.replace(" ", "")
        except AttributeError:
            pass
        data_list[39] = holiday_d_time
        #self.sheet.cell(row=index, column=40, value=holiday_d_time)  # 営業時間/定休日

        menu_list = [
            '駐車場',
            'クレジットカード',
            '座席',
            '用途',
            'メニュー',
            '特徴',
            'ポイント',
            'ここがすごい！',
            'メディア関連',
            '価格設定',
            'マルチアクセス'
        ]
        for col, menu in zip(range(41, 51+1), menu_list):
            #menu = self.sheet.cell(row=1, column=col).value
            try:
                data_list[col-1] = table_list[menu]
                #self.sheet.cell(row=index, column=col, value=table_list[menu])
            except KeyError:
                data_list[col-1] = None
                #self.sheet.cell(row=index, column=col, value=None)
        
        introduction_elm = soup.select_one('body > div.l-wrapper > div > div.l-contents_wrapper > main > div.l-shop_content > div:nth-child(2) > div > div > div.p-shop_introduction_content.js_toggle_content > p')
        introduction = introduction_elm.get_text() if introduction_elm != None else None
        data_list[51] = introduction
        #self.sheet.cell(row=index, column=52, value=introduction) #紹介文
        print(data_list)
        return data_list

    def calcLatLong(self, address:str):
        """
        using geocoding.jp API.
        this function return [latitude ,longitude].
        if callback error from API or return ConnectionError, will return ['取得失敗', '取得失敗'].       
        """
        url = 'http://www.geocoding.jp/api/'
        payload = {'q':address}
        try:
            html = requests.get(url, params=payload)
        except:
            time.sleep(60)
            try:
                html = requests.get(url, params=payload)
            except:
                return ['取得失敗', '取得失敗']
        soup = bs(html.content, 'lxml')
        if soup.find('error'):
            return ['取得失敗', '取得失敗']
        else:
            try:
                lat = soup.find('lat').string  #緯度
                long = soup.find('lng').string #経度
                return [lat, long]
            except AttributeError:
                return ['取得失敗', '取得失敗']
        
    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": 1,
            "青森県": 2,
            "岩手県": 3,
            "宮城県": 4,
            "秋田県": 5,
            "山形県": 6,
            "福島県": 7,
            "茨城県": 8,
            "栃木県": 9,
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        return code 

    
    def restart(self):
        """
        スクレイピング用ドライバおよびブラウザの再起動。
        """
        self.driver.delete_all_cookies()
        self.driver.quit()
        time.sleep(3)
        self.driver = webdriver.Chrome(
            'chromedriver.exe', options=self.options)
        self.wait = WebDriverWait(self.driver, 180)


class WriteWorkBook():
    book = px.Workbook()
    sheet = book.worksheets[0]

    def __init__(self, path):
        self.path = path
        #self.end_count = end_count
        self.__init_work_book()

    def __init_work_book(self):
        menu_list = [
            "エキテン",
            "新規リスト投入日",
            "ジャンル",
            "電話",
            "店舗名",
            "店舗名カナ",
            "料金プラン",
            "都道府県コード",
            "都道府県",
            "市区町村・番地",
            "住所フル",
            "店舗URL",
            "shopID",
            "URLその1",
            "URLその2",
            "URLその3",
            "パンくず",
            "キャッチ",
            "店舗公式",
            "未確認店舗",
            "評価点数",
            "口コミ数",
            "写真枚数",
            "アクセス",
            "ジャンル1",
            "ジャンル2",
            "ジャンル3",
            "ジャンル4",
            "早朝OK",
            "日祝OK",
            "夜間OK",
            "駐車場有",
            "ネット予約",
            "クーポン有",
            "カード可",
            "出張・宅配あり",
            "緯度",
            "経度",
            "アクセス／最寄駅",
            "営業時間／定休日",
            "駐車場",
            "クレジットカード",
            "座席",
            "用途",
            "メニュー",
            "特徴",
            "ポイント",
            "ここがすごい！",
            "メディア関連",
            "価格設定",
            "マルチアクセス",
            "紹介文"
        ]
        for col, menu in enumerate(menu_list):
            self.sheet.cell(row=1, column=col+1, value=menu)
        self.book.save(self.path)

    def write_data(self, data_list:list, index:int):
        """
        1プロセスあたりのスクレイピングデータを書き込む。\n
        data_list = [A,B,C, ....]
        """
        for col, data in zip(range(1, 52+1), data_list):
            try:
                self.sheet.cell(row=index, column=col, value=data)
            except IndexError:
                return False
        return True 

class Implementation():
    def __init__(self, path, area_list, junle):
        self.area_list = area_list
        self.junle = junle
        self.path = path
        self.manager = Manager()
        self.max_row_counter = self.manager.Value('i', 0)
        self.scrap_url_list = self.manager.list()
        self.end_count = self.manager.Value('i', 0)
        self.info_data_list = self.manager.list()
        self.writeBook = WriteWorkBook(self.path)
        self.search = ScrapingURL(path, self.max_row_counter, self.scrap_url_list)
        self.scrap = ScrapingInfomation(path, self.max_row_counter, self.scrap_url_list, self.info_data_list, self.end_count)
        self.scrap2 = ScrapingInfomation(path, self.max_row_counter, self.scrap_url_list, self.info_data_list, self.end_count)   
        self.scrap3 = ScrapingInfomation(path, self.max_row_counter, self.scrap_url_list, self.info_data_list, self.end_count)
        self.p = Pool()
        self.list1 = []
        self.list2 = []
        self.list3 = []
        self.search_sum = 1
        self.book_index = 2 #ワークシート書き込み行数の初期値
    
    def create_url_data_list(self):
        addCounter = 0
        addLength = len(self.scrap_url_list)
        while addCounter < addLength:
            try:
                self.list1.append(self.scrap_url_list.pop(0))
                self.list2.append(self.scrap_url_list.pop(0))
                self.list3.append(self.scrap_url_list.pop(0))
            except IndexError:
                #3つに渡せるほどない場合。このメソッドを即座に終了。
                break
            addCounter += 3

    def info_datas_writing(self):
        """
        共有メモリ上のリストを監視し、リストがemptyでない限り書き込みを続ける。
        """
        print("called!")
        length = len(self.info_data_list)
        #print("write_data_len : " + str(length))
        for i in range(0, length):
            #print(self.info_datas.empty())
            #print(self.info_datas[0])
            #data = self.info_datas[i]
            #print(data)
            try:
                self.writeBook.write_data(self.info_data_list.pop(0), self.book_index)
            except RemoteError:
                print(self.info_data_list)
                pass
                    #self.end_count += 1
            #self.writed_index += 1
            self.book_index += 1
        self.writeBook.book.save(self.writeBook.path)

    def call_counter_value(self):
        """
        呼び出し時点での進捗状況の値を返却する。[抽出済み件数, 検索結果数]
        """
        return self.end_count.value, self.search_sum

    def cancel(self):
        try:
            self.info_datas_writing()
            self.writeBook.book.save(self.writeBook.path)
            self.p.terminate()
        except:
            pass
        
    def run(self):
        #p = Pool()
        future = self.p.apply_async(self.search.search, args=([self.area_list]))
        scrap_flg = True
        search_flg = True

        while scrap_flg:
            self.search_sum = self.search.row_counter.value
            if len(self.scrap_url_list) > 0:
                self.create_url_data_list()
                result1 = self.p.apply_async(self.scrap.requestHTML, args=([self.list1]))
                result2 = self.p.apply_async(self.scrap2.requestHTML, args=([self.list2]))
                result3 = self.p.apply_async(self.scrap3.requestHTML, args=([self.list3]))
            
                doing = True
                async_result = [False, False, False]
                while doing:
                    self.search_sum = self.search.row_counter.value
                    if False not in async_result:
                        doing = False
                        if len(self.info_data_list) > 0:
                            self.info_datas_writing()
                        break

                    if result1.ready():
                        print("result1 end")
                        async_result[0] = True
                        self.list1.clear()
                        
                    if result2.ready():
                        print("result2 end")
                        async_result[1] = True
                        self.list2.clear()

                    if result3.ready():
                        print("result3 end")
                        async_result[2] = True
                        self.list3.clear()        

                self.search_sum = self.search.row_counter.value

                if future.ready():
                    future.get()
                    self.search_sum = self.search.row_counter.value
                    search_flg = False
                
                if (search_flg == False and
                    scrap_flg == True and 
                    self.search_sum == self.end_count.value):
                    scrap_flg = False
                    self.info_datas_writing()
                    self.writeBook.book.save(self.writeBook.path)
                    break

def resource_path(relative_path):#バイナリフィルのパスを提供
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

if __name__ == "__main__":
    test = Implementation('./multiprocess_run_test.xlsx', ['徳島県'], '全ジャンル抽出')
    test.run()