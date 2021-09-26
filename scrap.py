from PySimpleGUI.PySimpleGUI import No
import openpyxl as px
import re
import sys
import os
import time
from selenium import webdriver
from selenium.common.exceptions import InvalidSessionIdException, NoSuchElementException, WebDriverException
#from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup as bs
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import requests

class Scrap():
    book = px.Workbook()
    sheet = book.worksheets[0]

    def __init__(self, path):
        # init WebDriver
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_argument("enable-automation")
        self.options.add_argument("--headless")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument('--disable-extensions')
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-browser-side-navigation")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument('--ignore-certificate-errors')
        self.options.add_argument('--ignore-ssl-errors')
        #prefs = {"profile.default_content_setting_values.notifications": 2}
        #self.options.add_experimental_option("prefs", prefs)
        browser_path = resource_path('chrome-win/chrome.exe')
        self.options.binary_location = browser_path
        self.driver = webdriver.Chrome(executable_path='chromedriver.exe', options=self.options)
        self.sub_driver = webdriver.Chrome(executable_path='chromedriver.exe', options=self.options)
        self.wait_sub_driver = WebDriverWait(self.sub_driver, 180)
        self.path = path
        self.result_cnt = 2
        self.count = 0
        self.sheet_row = 1
        self.write_row = 1

    def book_init(self):
        col_list = [
            "エキテン",
            "新規リスト投入日",
            "ジャンル",
            "電話",
            "店舗名",
            "店舗名カナ",
            "料金プラン",
            "都道府県コード",
            "都道府県",
            "市区町村・番地",
            "住所フル",
            "店舗URL",
            "shopID",
            "URLその1",
            "URLその2",
            "URLその3",
            "パンくず",
            "キャッチ",
            "店舗公式",
            "未確認店舗",
            "評価点数",
            "口コミ数",
            "写真枚数",
            "アクセス",
            "ジャンル1",
            "ジャンル2",
            "ジャンル3",
            "ジャンル4",
            "早朝OK",
            "日祝OK",
            "夜間OK",
            "駐車場有",
            "ネット予約",
            "クーポン有",
            "カード可",
            "出張・宅配あり",
            "緯度",
            "経度",
            "アクセス／最寄駅",
            "営業時間／定休日",
            "駐車場",
            "クレジットカード",
            "座席",
            "用途",
            "メニュー",
            "特徴",
            "ポイント",
            "ここがすごい！",
            "メディア関連",
            "価格設定",
            "マルチアクセス",
            "紹介文"
        ]
        for col, menu in enumerate(col_list):
            self.sheet.cell(row=1, column=col+1, value=menu)
        self.book.save(self.path)

    def junle_list(self):
        list = [
            "リラク・ボディケア",
            "ヘアサロン・ネイル",
            "学習塾・予備校",
            "習い事・スクール",
            "歯科・矯正歯科",
            "医院・クリニック・ヘルスケア",
            "ショッピング",
            "お出かけ・レジャー",
            "リサイクル・中古買取り",
            "ペット・動物",
            "出張デリバリー・生活サービス",
            "住宅・不動産",
            "冠婚葬祭"
        ]
        return list

    # 全ジャンル抽出の場合
    def search(self, area):  # 検索と条件指定
        #driver_action = ActionChains(self.driver)
        # self.driver.implicitly_wait(5)
        self.sub_driver.get('https://www.ekiten.jp/')
        # wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#select_form_st_com")))
        # Max wait time(second):180s
        self.wait_sub_driver = WebDriverWait(self.sub_driver, 180)
        self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
        sr_box = self.sub_driver.find_element_by_id('select_form_st_com')
        sr_box.send_keys(area)
        sr_btn = self.sub_driver.find_element_by_css_selector(
            '#js_random_top > div > div > div > form > div > input')
        sr_btn.click()
        self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
        city_list = self.extraction_url(
            'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > div.grouped_list_body > ul > li > a', 'https://www.ekiten.jp')
        # print(city_list)
        result:str = self.sub_driver.find_element_by_css_selector(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.search_result_heading.u-mb10 > div.search_result_heading_sub > dl > div > dd').text
        result = result.replace(",", "")
        result = result.replace("件", "")
        self.result_cnt = int(result)
        self.count = 0
        for city in city_list:
            self.sub_driver.get(city)
            # wait.until(EC.visibility_of_all_elements_located)
            # print(city)
            # time.sleep(1)
            # WebDriverWait(self.driver, 180)#Max wait time(second):180s
            self.wait_sub_driver.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > a')))
            select = self.sub_driver.find_element_by_css_selector(
                'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > a').text

            # 区町村選択がない場合の処理系
            if select in '駅・バス停から探す ':
                # time.sleep(1)
                self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                junle_list = self.extraction_url(
                    'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                # print(junle_list)
                for junle in junle_list:
                    print(junle)
                    self.sub_driver.get(junle)
                    self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                    # time.sleep(1)
                    kategoli_list = self.extraction_url(
                        'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                    # print(kategoli_list)
                    for kategoli in kategoli_list:
                        print(kategoli)
                        self.sub_driver.get(kategoli)
                        self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                        # scrap URL process here
                        self.scrap_url()
                        self.count += 1
                    self.__sub_restart()
                self.__sub_restart()

            # 区町村選択がある場合の処理系
            else:
                self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                city_list2 = self.extraction_url(
                    'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(1) > ul > li:nth-child(3) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                for city2 in city_list2:
                    print(city2)
                    self.sub_driver.get(city2)
                    self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                    # time.sleep(1)
                    junle_list = self.extraction_url(
                        'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                    # print(junle_list)
                    for junle in junle_list:
                        print(junle)
                        self.sub_driver.get(junle)
                        self.wait_sub_driver.until(EC.visibility_of_all_elements_located)
                        # time.sleep(1)
                        kategoli_list = self.extraction_url(
                            'body > div.l-wrapper > div > div.l-contents_wrapper > div > nav > div:nth-child(2) > ul > li:nth-child(2) > div > div > div > div > div > ul > li > a', 'https://www.ekiten.jp')
                        # print(kategoli_list)
                        for kategoli in kategoli_list:
                            print(kategoli)
                            self.sub_driver.get(kategoli)
                            self.wait_sub_driver.until(
                                EC.visibility_of_all_elements_located)
                            self.scrap_url()
                            self.count += 1
                            # scrap URL process here
                        self.__sub_restart()
                self.__sub_restart()
            self.__sub_restart()
        self.__sub_restart()

    def __sub_restart(self):
        self.sub_driver.delete_all_cookies()
        self.sub_driver.quit()
        time.sleep(3)
        self.sub_driver = webdriver.Chrome(
            'chromedriver.exe', options=self.options)
        self.wait_sub_driver = WebDriverWait(self.sub_driver, 180)
    
    def restart(self):
        self.driver.delete_all_cookies()
        self.driver.quit()
        time.sleep(3)
        self.driver = webdriver.Chrome(
            'chromedriver.exe', options=self.options)
        self.wait = WebDriverWait(self.driver, 180)

    def scrap_url(self):
        while True:
            html = self.sub_driver.page_source
            soup = bs(html, 'lxml')
            try:
                a_tags = soup.select('.p-shop_box_head_title_body > a')
                # print(a_tags)
                self.write_url(a_tags)
                next_btn = self.sub_driver.find_element_by_css_selector(
                    'div.p-pagination_next > a')
                next_btn.click()
            except NoSuchElementException:
                break

    def info_scrap(self, url, index):
        self.sheet.cell(row=index, column=1, value=None)  # A列
        self.sheet.cell(row=index, column=2, value=None)  # B列
        self.driver.get(url)
        html = self.driver.page_source
        soup = bs(html, 'lxml')
        junle_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.layout_media.p-topic_path_container > div.layout_media_wide > div > a:nth-child(3)'
        )
        junle = junle_elm.get_text() if junle_elm != None else None
        self.sheet.cell(row=index, column=3, value=junle)  # ジャンル
        tel_elm = soup.select_one(
            'body > div:nth-child(12) > div > div.p-tel_modal_phone_number > div > div > p'
        )
        tel = tel_elm.get_text() if tel_elm != None else None
        self.sheet.cell(row=index, column=4, value=tel)  # TEL
        # table info scraping
        table_col = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.p-shop_content_container.p-shop_content_container_relative > table > tbody > tr > th'
        )
        table_info = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.p-shop_content_container.p-shop_content_container_relative > table > tbody > tr >td'
        )
        table_list = {}
        for menu, info in zip(table_col, table_info):
            try:
                menu = menu.get_text()
                info = info.get_text()
            except AttributeError:
                break
            info = info.replace(" ", "")
            info = info.replace("　", "")
            info = info.replace("\n", "")
            info = info.replace("地図で場所を見るGoogleマップで見る", "")
            table_list[menu] = info
        self.sheet.cell(row=index, column=5, value=table_list['店舗名'])
        store_kana_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div > div.p-shop_header_name_container > div > span'
            )
        store_kana = store_kana_elm.get_text() if store_kana_elm != None else None
        self.sheet.cell(row=index, column=6, value=store_kana)
        all_addresses = table_list['住所']
        pref_obj = re.search('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_addresses)
        pref = pref_obj.group()
        self.sheet.cell(row=index, column=8,
                        value=self.call_jis_code(pref))  # JISコード
        self.sheet.cell(row=index, column=9, value=pref)  # 都道府県名
        # 都道府県と市区町村を分離
        muni = re.split('東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_addresses)
        self.sheet.cell(row=index, column=10, value=muni[1])  # 市区町村番地
        self.sheet.cell(row=index, column=11, value=all_addresses)  # フル住所
        self.sheet.cell(row=index, column=12, value=url)  # 店舗URL
        shop_id_string = re.search(r"shop_\d{0,}", url).group()
        shop_id = re.search(r"\d{1,}", shop_id_string).group()
        self.sheet.cell(row=index, column=13, value=shop_id)  # shopID
        try:
            url_list = re.findall(
                r"https?://[\w!\?/\+\-_~=;\.,\*&@#\$%\(\)'\[\]]+", table_list['URL'])
        except KeyError:  # URLがない時
            url_list = [None, None, None]
        for i in range(3):
            try:
                self.sheet.cell(row=index, column=14+i,
                                value=url_list[i])  # URLその１～３
            except IndexError:
                self.sheet.cell(row=index, column=14+i, value=None)
        pankuzu_header = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.layout_media.p-topic_path_container > div.layout_media_wide > div').get_text()
        pankuzu = pankuzu_header.replace('\n', " > ")
        self.sheet.cell(row=index, column=17,
                        value=pankuzu.strip(" > "))  # パンくずヘッダー
        
        catch_copy_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.p-shop_header_catch_container > p'
            )
        catch_copy = catch_copy_elm.get_text() if catch_copy_elm != None else None
        self.sheet.cell(row=index, column=18, value=catch_copy)
        official = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.p-shop_header_catch_container > div > span')
        official_judge = '●' if official != None else None
        self.sheet.cell(row=index, column=19, value=official_judge)
        self.sheet.cell(row=index, column=20, value=None)  # 未確認店舗
        store_score_tag = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_rating > div > div.rating_stars_num.tooltip > span'
            )
        if store_score_tag == None:
            score = None  # 評価点数
        else:
            score = float(store_score_tag.get_text())
        self.sheet.cell(row=index, column=21, value=score)
        review_count = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_info > div.p-shop_header_info_review > div > span.icon_wrapper_text > a')
        if review_count in (None, '0'):
            review_count = 0
        else:
            review_count = int(review_count.get_text())
        self.sheet.cell(row=index, column=22, value=review_count)  # 口コミ数
        photo_tag = soup.select(
            'body > div.l-wrapper > div > div.l-contents_wrapper > main > div.l-shop_content > div:nth-child(3) > div.grid.space15.vertical_space15.js_photo_gallery > div > a > img')
        photo_cnt = len(photo_tag)  # 写真枚数
        self.sheet.cell(row=index, column=23, value=photo_cnt)
        access_elm = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_access > div:nth-child(1) > div')
        access = access_elm.get_text() if access_elm != None else None
        self.sheet.cell(row=index, column=24, value=access)  # アクセス
        
        junle1_elm = soup.select_one('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > ul.p-shop_header_genre > li > a')
        junle1 = junle1_elm.get_text() if junle1_elm != None else None
        self.sheet.cell(row=index, column=25, value=junle1)  # ジャンル１
        
        junle2_4 = soup.select('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > ul.p-shop_header_genre > li > span')
        #ジャンル2~4
        for c, junle in enumerate(junle2_4):
            if junle != None:
                self.sheet.cell(row=index, column=26+c, value=junle.get_text())
        
        
        features = soup.select('body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > ul.p-shop_header_tag_list.tag_group > li')
        for feature in features:
            if feature == None:  # 特徴タグがない時処理を行わない
                break
            
            for col in range(29, 36+1): 
                if self.sheet.cell(row=1, column=col).value in feature.get_text():
                    self.sheet.cell(row=index, column=29+i, value="●")  # 店舗の特徴
                    break  # 見つかったら小ループ抜けて次の特徴へ
        
        #緯度・経度
        latlong = self.calcLatLong(all_addresses)
        self.sheet.cell(row=index, column=37, value=latlong[0])
        self.sheet.cell(row=index, column=38, value=latlong[1])

        moyorieki = None
        self.sheet.cell(row=index, column=39, value=moyorieki)  # アクセス/最寄り駅
        holiday_d_time_elm = soup.select_one(
            'body > div.l-wrapper > div > div.l-top_contents > div.p-shop_header > div.layout_media.p-shop_header_inner > div.layout_media_wide.p-shop_header_main > div.layout_media.p-shop_header_main_inner > div.layout_media_wide.p-shop_header_main_content > div:nth-child(1) > div > div.p-shop_header_access > div:nth-child(3) > div'
        )
        holiday_d_time = holiday_d_time_elm.get_text() if holiday_d_time_elm != None else None
        try:
            holiday_d_time = holiday_d_time.replace("\n", "/")
            holiday_d_time = holiday_d_time.replace(" ", "")
        except AttributeError:
            pass
        self.sheet.cell(row=index, column=40, value=holiday_d_time)  # 営業時間/定休日

        for col in range(41, 51+1):
            menu = self.sheet.cell(row=1, column=col).value
            try:
                self.sheet.cell(row=index, column=col, value=table_list[menu])
            except KeyError:
                self.sheet.cell(row=index, column=col, value=None)
        
        introduction_elm = soup.select_one('body > div.l-wrapper > div > div.l-contents_wrapper > main > div.l-shop_content > div:nth-child(2) > div > div > div.p-shop_introduction_content.js_toggle_content > p')
        introduction = introduction_elm.get_text() if introduction_elm != None else None
        self.sheet.cell(row=index, column=52, value=introduction) #紹介文

    def calcLatLong(self, address:str):
        """
        using geocoding.jp API.
        this function return [latitude ,longitude].
        if callback error from API, will return ['取得失敗', '取得失敗'].       
        """
        url = 'http://www.geocoding.jp/api/'
        payload = {'q':address}
        try:
            html = requests.get(url, params=payload)
        except ConnectionError:
            time.sleep(60)
            try:
                html = requests.get(url, params=payload)
            except ConnectionError:
                return ['取得失敗', '取得失敗']
        soup = bs(html.content, 'lxml')
        if soup.find('error'):
            return ['取得失敗', '取得失敗']
        else:
            try:
                lat = soup.find('lat').string  #緯度
                long = soup.find('lng').string #経度
                return [lat, long]
            except AttributeError:
                return ['取得失敗', '取得失敗']
        
    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": 1,
            "青森県": 2,
            "岩手県": 3,
            "宮城県": 4,
            "秋田県": 5,
            "山形県": 6,
            "福島県": 7,
            "茨城県": 8,
            "栃木県": 9,
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        return code

    def write_url(self, a_tag_list):
        """
        抽出URLの書き込み(シート12列目)
        """
        url_list = []
        for a in a_tag_list:
            url_list.append("https://www.ekiten.jp" + a.get('href'))

        for url in url_list:
            try:
                row = self.sheet.max_row + 1
            except RuntimeError:
                row = self.write_row + 1
            isWrite = True
            try:
                preload = self.sheet.max_row
            except RuntimeError:
                preload = self.write_row
            for r in range(2, preload):  # 重複チェック用for loop
                pre_url = self.sheet.cell(row=r, column=12).value
                if pre_url == url:
                    isWrite = False
                    break
            if isWrite:
                print(url)
                self.sheet.cell(row=row, column=12, value=url)
                self.write_row += 1
        try:
            self.sheet_row = self.sheet.max_row
        except RuntimeError:
            self.sheet_row = self.write_row
        #self.book.save(self.path)

    def extraction_url(self, selector, pre_url):
        html = self.sub_driver.page_source
        soup = bs(html, 'lxml')
        s = soup.select(selector)
        # #print(s)
        url_list = self.return_url(s, pre_url)
        return url_list

    def return_url(self, a_tag_list, pre_url):
        url_list = []
        for a in a_tag_list:
            url_list.append(pre_url+a.get('href'))
        return url_list


def resource_path(relative_path):  # バイナリフィルのパスを提供
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


if __name__ == "__main__":
    """
    test_url = [
        "https://www.ekiten.jp/shop_88106804/",
        "https://www.ekiten.jp/shop_79608272/",
        "https://www.ekiten.jp/shop_70897950/",
        "https://www.ekiten.jp/shop_70296422/",
        "https://www.ekiten.jp/shop_58067946/",
        "https://www.ekiten.jp/shop_51517853/",
        "https://www.ekiten.jp/shop_63474300/",
        "https://www.ekiten.jp/shop_16443976/",
        "https://www.ekiten.jp/shop_27214855/",
        "https://www.ekiten.jp/shop_49566534/",
        "https://www.ekiten.jp/shop_41288052/",
        "https://www.ekiten.jp/shop_13378029/",
        "https://www.ekiten.jp/shop_65698495/",
        "https://www.ekiten.jp/shop_89888651/",
        #"https://www.ekiten.jp/shop_50246627/",
        "https://www.ekiten.jp/shop_27904560/",
        "https://www.ekiten.jp/shop_43582014/",
        "https://www.ekiten.jp/shop_54599516/",
        "https://www.ekiten.jp/shop_20556262/",
        "https://www.ekiten.jp/shop_65198684/",
        "https://www.ekiten.jp/shop_36161923/",
        "https://www.ekiten.jp/shop_94297307/",
        "https://www.ekiten.jp/shop_84159649/",
        "https://www.ekiten.jp/shop_65817519/",
        "https://www.ekiten.jp/shop_69489840/",
        "https://www.ekiten.jp/shop_26687442/",
        "https://www.ekiten.jp/shop_22073739/",
        "https://www.ekiten.jp/shop_61620533/",
        "https://www.ekiten.jp/shop_95295090/",
    ]
    """
    scraping = Scrap('./run_test.xlsx')
    scraping.book_init()
    scraping.search('徳島県')
    for r in range(2, scraping.sheet.max_row+1):
        url = scraping.sheet.cell(row=r, column=13).value
        #print(str(r) + ".scraping at:" + url)
        scraping.info_scrap(url, r)
    # Test Run
    # print("compleate!!")
    scraping.driver.quit()
