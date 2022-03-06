from __future__ import annotations
from typing import Iterator, Optional, Any, Final as const

import openpyxl as pyxl
from abc import ABC, ABCMeta, abstractmethod
import os

COLUMN_MENUS:const[dict] = {
    "store_big_junle" : "ジャンル",
    "store_tel" : "電話",
    "store_name" : "店舗名",
    "str_name_kana" : "店舗名カナ",
    "price_plan" : "料金プラン",
    "pref_code" : "都道府県コード",
    "pref" : "都道府県",
    "city" : "市区町村・番地",
    "full_address" : "住所フル",
    "store_link" : "店舗URL",
    "shop_id" : "shopID",
    "st_home_page1" : "URLその1",
    "st_home_page2" : "URLその2",
    "st_home_page3" : "URLその3",
    "pankuzu" : "パンくず",
    "catch_cp" : "キャッチ",
    "is_official" : "店舗公式",
    "evaluation_score" : "評価点",
    "review_count" : "口コミ数",
    "image_count" : "写真枚数",
    "access_info" : "アクセス",
    "junle1" : "ジャンル1",
    "junle2" : "ジャンル2",
    "junle3" : "ジャンル3",
    "junle4" : "ジャンル4",
    "can_early_morning" : "早朝OK",
    "can_date_celebration" : "日祝OK",
    "can_night" : "夜間OK",
    "has_park" : "駐車場あり",
    "can_rev_net" : "ネット予約",
    "has_coupon" : "クーポンあり",
    "can_card" : "カード可",
    "can_delivery" : "出張・宅配あり",
    "latitude" : "緯度",
    "longitude" : "経度",
    "closest_station" : "最寄り駅",
    "business_hours" : "営業時間/定休日",
    "park_info" : "駐車場",
    "credit_info" : "クレジットカード",
    "seat_info" : "座席",
    "use_case" : "用途",
    "menu" : "メニュー",
    "feature" : "特徴",
    "point" : "ポイント",
    "here_is_great" : "ここがすごい",
    "media_related": "メディア関連",
    "pricing" : "価格設定",
    "multi_access" : "マルチアクセス",
    "introduce" : "紹介文",
}

class AbsWorkBook(object, metaclass=ABCMeta):
    
    """[summary]\n 
    指定のExcelファイルを生成編集する。その抽象クラス。
    """
    
    colm_menu:dict[str, str] = COLUMN_MENUS

    
    def __init__(self, save_file_path:str, crawl_file_path_list:list[str]) -> None:
        
        self.file_path:const[str] = save_file_path #最終的に保存するファイルのパス
        self.distributed_files_list:const[list[str]] = crawl_file_path_list[1:len(crawl_file_path_list)] #分割したファイルのパスリスト
        print(self.distributed_files_list)
        try:
            self.book = pyxl.load_workbook(crawl_file_path_list[0]) #分割したファイルの最初のファイルを読み込む
        except FileNotFoundError:
            raise FileNotFoundError("一時保存ファイルが見つからないため、xlsxファイルを作成できませんでした。")
        self.worksheet  = self.book.worksheets[0]
        
    
    @abstractmethod
    def col_menulocalize(self) -> None:
        pass
    
    @abstractmethod
    def create_crawl_workbook(self) -> None:
        """_summary_\n
        分散クロールの一時保存ファイルの内容を結合し、1つのExcelファイルとする。
        """
        pass
    
    @abstractmethod
    def remove_temp_file(self) -> None:
        """_summary_\n
        一時保存ファイルを削除する。
        """
        pass
        
    
    def save(self):
        self.book.save(self.file_path)
        

class WorkBook(AbsWorkBook):
    """[summary]\n
    各クローラーの結果を統合し、1つのExcelファイルにする。
    """
    
    def __init__(self, save_file_path: str, crawled_file_path:list[str]) -> None:
        super().__init__(save_file_path, crawled_file_path)
    
    def create_crawl_workbook(self) -> None:
        """_summary_\n
        すべてのクロール結果を統合し、1つのExcelファイルを作成する。
        """
        #TODO: 分散クロールの一時保存ファイルの内容を結合し、1つのExcelファイルとする。
        
        for path in self.distributed_files_list:
            book = self.__loadBook(path)
            self.__sheet_combined(book)
            print("{}番目のファイルを統合しました。".format(self.distributed_files_list.index(path)))
        
        
    def __sheet_combined(self, book:pyxl.Workbook) -> None:
        """_summary_\n 
        ワークシートをコピーする。
        """
        origin_sheet_rows: Iterator[Any] = book.worksheets[0].rows #type: ignore
        write_start_row: int = self.worksheet.max_row + 1 #type: ignore
        for i, row in enumerate(origin_sheet_rows):
            if i == 0: #ヘッダー行はコピーしない
                continue
            self.__cell_copy(write_start_row + i-1, row)
    
    
    def __cell_copy(self, write_row: int,row_values: Iterator[Any]):
        for i, value in enumerate(row_values):
            self.worksheet.cell(row=write_row, column=i + 1, value=value.value)
            print("{}行目{}列目をコピーしました。".format(write_row + i, i + 1))
            
    
    def __loadBook(self, filename: str) -> pyxl.Workbook:
        book:pyxl.Workbook = pyxl.load_workbook(filename)
        return book
        
    
    def remove_temp_file(self) -> None:
        """_summary_\n 
        一時保存ファイルを削除する。\n
        caution: 分散クロールの一時保存ファイルを削除するため、save()を実行してから実行すること。
        """
        for path in self.distributed_files_list:
            os.remove(path)
        
    def col_menulocalize(self):
        """\n
        クロール後のExcelファイルの、
        先頭行の変数で記載された項目を日本語化する。
        """
        max_colum: Optional[int] = self.worksheet.max_column
        assert max_colum is not None
        for col in range(1, max_colum + 1):
            original:Any = self.worksheet.cell(row=1, column=col).value
            self.worksheet.cell(row=1, column=col).value = self.colm_menu[original] #type: ignore

#test call
if __name__ == '__main__':
    #API TEST CALL
    combine_list: const[list[str]]
    combine_list = ["save_test copy 2.xlsx", "save_test copy 3.xlsx", "save_test copy 4.xlsx", "save_test copy.xlsx"]
    test = WorkBook("save_test-API.xlsx", combine_list)
    test.create_crawl_workbook()
    test.save()


