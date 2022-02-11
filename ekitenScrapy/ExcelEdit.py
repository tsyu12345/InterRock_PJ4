import openpyxl as pyxl
from abc import ABC, ABCMeta, abstractmethod


class AbsExcelEdit(object, metaclass=ABCMeta):
    
    """[summary]\n 
    指定のExcelファイルを編集する。その抽象クラス。
    """
    
    def __init__(self, file_path:str) -> None:
        
        self.file_path = file_path
        
        self.book = pyxl.load_workbook(self.file_path)
        self.worksheet = self.book.worksheets[0]
        
        self.colm_menu:dict = {
            "store_big_junle" : "ジャンル",
            "store_tel" : "電話",
            "store_name" : "店舗名",
            "str_name_kana" : "店舗名カナ",
            "price_plan" : "料金プラン",
            "pref_code" : "都道府県コード",
            "pref" : "都道府県",
            "city" : "市区町村・番地",
            "full_address" : "住所フル",
            "store_link" : "店舗URL",
            "shop_id" : "shopID",
            "st_home_page1" : "URLその1",
            "st_home_page2" : "URLその2",
            "st_home_page3" : "URLその3",
            "pankuzu" : "パンくず",
            "catch_cp" : "キャッチ",
            "is_official" : "店舗公式",
            "evaluation_score" : "評価点",
            "review_count" : "口コミ数",
            "image_count" : "写真枚数",
            "access_info" : "アクセス",
            "junle1" : "ジャンル1",
            "junle2" : "ジャンル2",
            "junle3" : "ジャンル3",
            "junle4" : "ジャンル4",
            "can_early_morning" : "早朝OK",
            "can_date_celebration" : "日祝OK",
            "can_night" : "夜間OK",
            "has_park" : "駐車場あり",
            "can_rev_net" : "ネット予約",
            "has_coupon" : "クーポンあり",
            "can_card" : "カード可",
            "can_delivery" : "出張・宅配あり",
            "latitude" : "緯度",
            "longitude" : "経度",
            "closest_station" : "最寄り駅",
            "business_hours" : "営業時間/定休日",
            "park_info" : "駐車場",
            "credit_info" : "クレジットカード",
            "seat_info" : "座席",
            "use_case" : "用途",
            "menu" : "メニュー",
            "feature" : "特徴",
            "point" : "ポイント",
            "here_is_great" : "ここがすごい",
            "media_related": "メディア関連",
            "pricing" : "価格設定",
            "multi_access" : "マルチアクセス",
            "introduce" : "紹介文",
        }
    
    @abstractmethod
    def col_menulocalize(self):
        pass
    
    
    def save(self):
        self.book.save(self.file_path)
        

class ExcelEdit(AbsExcelEdit):
    """[summary]\n
    指定のExcelファイルを編集する。その実装クラス。\n
    主目的は指定行のセル値を日本語化。
    """
    
    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        
    def col_menulocalize(self):
        """\n
        クロール後のExcelファイルの、
        先頭行の変数で記載された項目を日本語化する。
        """
        for col in range(1, self.worksheet.max_column + 1):
            original:str = self.worksheet.cell(row=1, column=col).value
            self.worksheet.cell(row=1, column=col).value = self.colm_menu[original]

#test call
if __name__ == '__main__':
    test = ExcelEdit('test.xlsx')
    test.col_menulocalize()
    test.save()
    


